# src/matrix/export_excel.py
from __future__ import annotations
from pathlib import Path
import json
from typing import List, Dict, Any, Union
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

PathLike = Union[str, Path]

PREFERRED_ORDER = [
    "id", "label", "category", "modality",
    "quote", "section", "page_start", "page_end",
    "confidence", "source", "doc_name",
]

def _coerce_int(v: Any) -> Any:
    """Try to convert to int, return original if fails."""
    try:
        if v is None or v == "":
            return None
        return int(v)
    except Exception:
        return v

def _union_columns(rows: List[Dict[str, Any]]) -> List[str]:
    """Get all unique columns from rows, with preferred order first."""
    keys = set()
    for r in rows:
        keys.update(r.keys())
    
    # Preferred columns first (if present), then extras alphabetically
    extras = sorted([k for k in keys if k not in PREFERRED_ORDER])
    return [k for k in PREFERRED_ORDER if k in keys] + extras

def save_excel(reqs: List[Dict[str, Any]], path: PathLike) -> Path:
    """
    Write requirements to Excel using openpyxl with nice formatting.
    - Capitalized headers
    - Excel Table with banded rows
    - Auto-sized columns
    """
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.styles import Font, Alignment
    
    out_path = Path(path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    
    if not reqs:
        # Create empty workbook if no data
        wb = Workbook()
        wb.save(out_path)
        return out_path
    
    # Get all columns from all requirements
    columns = _union_columns(reqs)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance Matrix"
    
    # Write header with capitalized names
    capitalized_headers = [col.replace("_", " ").title() for col in columns]
    ws.append(capitalized_headers)
    
    # Style header row
    for cell in ws[1]:
        cell.font = Font(bold=True, size=11, color="0000FF")
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Write data rows
    for req in reqs:
        row = []
        for col in columns:
            val = req.get(col)
            
            # Defensive serialization for non-scalars
            if isinstance(val, (dict, list, set, tuple)):
                val = json.dumps(val, ensure_ascii=False)
            
            # Try to coerce page numbers to integers
            if col in ("page_start", "page_end", "page"):
                val = _coerce_int(val)
            
            row.append(val)
        ws.append(row)
    
    # Auto-size columns (with max width limit)
    for idx, col_name in enumerate(columns, start=1):
        max_len = len(capitalized_headers[idx-1])
        col_letter = get_column_letter(idx)
        for cell in ws[col_letter]:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                max_len = max(max_len, cell_len)
        # Set width with reasonable limits
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
    
    # Create Excel Table with banded rows
    last_row = len(reqs) + 1  # +1 for header
    last_col = get_column_letter(len(columns))
    table_ref = f"A1:{last_col}{last_row}"
    
    tab = Table(displayName="ComplianceMatrix", ref=table_ref)
    
    # Style: Medium blue banded rows (TableStyleMedium2)
    style = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,  # Banded rows
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Set default row height for better readability
    ws.row_dimensions[1].height = 20
    
    wb.save(out_path)
    return out_path